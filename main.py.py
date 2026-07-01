#Muhammad Farzeen Khan Tareen
#23i-0721
#CS-B

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Zones and Capacity
zones = ["Z1","Z2","Z3","Z4","Z5","Z6","Z7","Z8"]

zone_capacity = { "Z1":120,"Z2":80,"Z3":80,"Z4":80,"Z5":80,"Z6":60,"Z7":150,"Z8":100 }

zone_names = {
    "Z1":"Heavy Item Floor Storage",
    "Z2":"Standard Rack Storage",
    "Z3":"Fragile Item Shelf",
    "Z4":"Temperature Controlled Storage",
    "Z5":"Hazardous Material Storage",
    "Z6":"Fast-Moving Product Area (Near Exit)",
    "Z7":"Bulk Dry Storage",
    "Z8":"Refrigerated Loading Dock"
}

# Product Data

products = [
("Glass Bottles",20,"Beverage",True,False,False,"Medium"),
("Frozen Meat",30,"Food",False,False,True,"High"),
("Cleaning Acid",10,"Chemical",False,True,False,"Low"),
("Rice Bags",50,"Grocery",False,False,False,"High"),
("Ceramic Plates",15,"Kitchenware",True,False,False,"Medium"),
("Ice Cream",25,"Food",False,False,True,"High"),
("Detergent",12,"Chemical",False,True,False,"Medium"),
("Chips Carton",8,"Snacks",False,False,False,"High"),
("Olive Oil Bottles",18,"Grocery",True,False,False,"Medium"),
("Industrial Bleach",22,"Chemical",False,True,False,"Low"),
("Yogurt Cartons",14,"Food",False,False,True,"High"),
("Flour Bags",45,"Grocery",False,False,False,"High"),
("Wine Bottles",16,"Beverage",True,False,False,"Medium"),
("Paint Cans",28,"Chemical",False,True,False,"Low"),
("Biscuit Boxes",6,"Snacks",False,False,False,"High"),
("Motor Oil",35,"Chemical",False,True,False,"Low"),
("Frozen Fish",20,"Food",False,False,True,"High"),
("Bubble Wrap Rolls",10,"Packaging",True,False,False,"Medium"),
("Wheat Sacks",40,"Grocery",False,False,False,"High"),
("Hand Sanitizer",8,"Chemical",True,False,False,"Medium")
]


# Fitness Function
def fitness(chromosome):

    penalty = 0
    zone_weight = {z:0 for z in zones}

    for i,zone in enumerate(chromosome):
        zone_weight[zone] += products[i][1]

    for z in zones:
        if zone_weight[z] > zone_capacity[z]:
            excess = zone_weight[z] - zone_capacity[z]
            penalty += excess * 2

    for i,zone in enumerate(chromosome):

        name,weight,cat,frag,haz,temp,demand = products[i]

        if frag and zone!="Z3":
            penalty += 8

        if haz and zone!="Z5":
            penalty += 10

        if temp and zone not in ["Z4","Z8"]:
            penalty += 9

        if demand=="High" and zone!="Z6":
            penalty += 5

        if weight>40 and zone!="Z1":
            penalty += 4

        if zone=="Z8":
            if not(temp and demand=="High"):
                penalty += 12

    # category compatibility
    for i in range(len(products)):
        for j in range(i+1,len(products)):
            if products[i][2] == products[j][2]:
                if chromosome[i] != chromosome[j]:
                    penalty += 3

    # food vs chemical incompatibility
    for i in range(len(products)):
        for j in range(i+1,len(products)):

            cat_i = products[i][2]
            cat_j = products[j][2]

            if (cat_i=="Food" and cat_j=="Chemical") or (cat_i=="Chemical" and cat_j=="Food"):
                if chromosome[i] == chromosome[j]:
                    penalty += 15

    return penalty



# Random Chromosome
def random_chromosome():
    return [random.choice(zones) for _ in range(len(products))]

def initial_population(size):
    return [random_chromosome() for _ in range(size)]


# Tournament Selection

def selection(pop):

    tournament = random.sample(pop,5)

    best = tournament[0]
    best_score = fitness(best)

    for c in tournament[1:]:

        score = fitness(c)

        if score < best_score:
            best = c
            best_score = score

    return best


# Crossover

def crossover(p1,p2):

    point = random.randint(1,len(products)-1)

    c1 = p1[:point] + p2[point:]
    c2 = p2[:point] + p1[point:]

    return c1,c2


# Mutation

def mutation(chromosome,rate=0.05):

    for i in range(len(chromosome)):

        if random.random() < rate:
            chromosome[i] = random.choice(zones)

    return chromosome



# Genetic Algorithm

def genetic_algorithm():

    population_size = 100
    generations = 300

    pop = initial_population(population_size)

    best = pop[0]

    for g in range(generations):

        pop.sort(key=fitness)
        best = pop[0]

        new_pop = [best.copy()]  # elitism

        while len(new_pop) < population_size:

            p1 = selection(pop)
            p2 = selection(pop)

            c1,c2 = crossover(p1,p2)

            c1 = mutation(c1)
            c2 = mutation(c2)

            new_pop.append(c1)

            if len(new_pop) < population_size:
                new_pop.append(c2)

        pop = new_pop

        print("Generation",g+1,"Best Fitness",fitness(best))

        if fitness(best)==0:
            break

    return best,g+1


# Show Storage Plan

def show_solution(solution):

    zone_map = {z:[] for z in zones}

    for i,zone in enumerate(solution):
        zone_map[zone].append(products[i][0])

    print("\nOptimal Storage Plan\n")

    for z in zones:

        if zone_map[z]:
            print(z+":",", ".join(zone_map[z]))
        else:
            print(z+": empty")


# Violation Report

def violation_report(chromosome):

    violations = []

    zone_weight = {z:0 for z in zones}

    for i,z in enumerate(chromosome):
        zone_weight[z] += products[i][1]

    for z in zones:
        if zone_weight[z] > zone_capacity[z]:
            violations.append("Weight capacity exceeded in "+z)

    for i,zone in enumerate(chromosome):

        name,weight,cat,frag,haz,temp,demand = products[i]

        if frag and zone!="Z3":
            violations.append(name+" fragile not in Z3")

        if haz and zone!="Z5":
            violations.append(name+" hazardous not in Z5")

        if temp and zone not in ["Z4","Z8"]:
            violations.append(name+" temperature storage violation")

        if demand=="High" and zone!="Z6":
            violations.append(name+" high-demand item not in Z6")

        if weight>40 and zone!="Z1":
            violations.append(name+" heavy item not in Z1")

        if zone=="Z8":
            if not(temp and demand=="High"):
                violations.append(name+" not eligible for Z8")

    # food + chemical conflict reporting
    for i in range(len(products)):
        for j in range(i+1,len(products)):

            cat_i = products[i][2]
            cat_j = products[j][2]

            if (cat_i=="Food" and cat_j=="Chemical") or (cat_i=="Chemical" and cat_j=="Food"):

                if chromosome[i] == chromosome[j]:

                    p1 = products[i][0]
                    p2 = products[j][0]
                    z  = chromosome[i]

                    violations.append(p1+" and "+p2+" food-chemical conflict in "+z)

    return violations


# Bonus: Export Results to Excel

def export_to_excel(solution, best_fitness, generations_run, filename="warehouse_plan.xlsx"):

    wb = openpyxl.Workbook()

    zone_colors = {
        "Z1":"FCE4D6",  # light orange  Heavy Floor
        "Z2":"EDEDED",  # light grey    Standard Rack
        "Z3":"E2EFDA",  # light green   Fragile Shelf
        "Z4":"DDEBF7",  # light blue    Temp Controlled
        "Z5":"FFCCCC",  # light red     Hazardous
        "Z6":"FFF2CC",  # light yellow  Fast-Moving
        "Z7":"F4CCFF",  # light purple  Bulk Dry
        "Z8":"D9E1F2"   # soft blue     Refrigerated Dock
    }

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    center = Alignment(horizontal="center", vertical="center")

    # SHEET 1
    ws1 = wb.active
    ws1.title = "Product Assignments"

    headers1 = ["#", "Product", "Weight (kg)", "Category","Fragile", "Hazardous", "Temp Controlled","Demand", "Assigned Zone", "Zone Description"]

    for col, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, zone in enumerate(solution):
        name, weight, cat, frag, haz, temp, demand = products[i]
        row_data = [i+1, name, weight, cat,
                    "Yes" if frag  else "No",
                    "Yes" if haz   else "No",
                    "Yes" if temp  else "No",
                    demand, zone, zone_names[zone]]
        row_num = i + 2
        for col, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=zone_colors[zone])
            cell.alignment = center

    # Auto-fit column widths for sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max_len + 4

    # SHEET 2
    ws2 = wb.create_sheet("Zone Summary")

    headers2 = ["Zone", "Zone Description", "Capacity (kg)",
                "Total Weight (kg)", "Remaining (kg)", "# Products", "Products Stored"]

    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    zone_weight = {z: 0 for z in zones}
    zone_products = {z: [] for z in zones}

    for i, zone in enumerate(solution):
        zone_weight[zone]   += products[i][1]
        zone_products[zone].append(products[i][0])

    for row_num, z in enumerate(zones, start=2):
        total = zone_weight[z]
        cap = zone_capacity[z]
        remain = cap - total
        prods = ", ".join(zone_products[z]) if zone_products[z] else "empty"
        count = len(zone_products[z])

        row_data = [z, zone_names[z], cap, total, remain, count, prods]

        for col, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_num, column=col, value=val)
            cell.alignment = center
            # Red if over capacity, normal zone color otherwise
            if col == 5 and remain < 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")
                cell.font = Font(bold=True, color="FFFFFF")
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor=zone_colors[z])

    for col in ws2.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max_len + 4

    # SHEET 3
    ws3 = wb.create_sheet("Violation Report")

    # Title row
    title_cell = ws3.cell(row=1, column=1, value="Warehouse Constraint Violation Report")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="C00000")
    title_cell.alignment = center
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 25

    # Summary info
    ws3.cell(row=3, column=1, value="Best Fitness Score").font = Font(bold=True)
    ws3.cell(row=3, column=2, value=best_fitness)
    ws3.cell(row=4, column=1, value="Generations Run").font    = Font(bold=True)
    ws3.cell(row=4, column=2, value=generations_run)

    # Violations header
    vh1= ws3.cell(row=6, column=1, value="#")
    vh1.font= Font(bold=True)
    vh2= ws3.cell(row=6, column=2, value="Violation Description")
    vh2.font= Font(bold=True)

    violations = violation_report(solution)

    if violations:
        for idx, v in enumerate(violations, start=1):
            ws3.cell(row=6+idx, column=1, value=idx)
            cell= ws3.cell(row=6+idx, column=2, value=v)
            cell.fill = PatternFill(fill_type="solid", fgColor="FFE0E0")
    else:
        ok_cell= ws3.cell(row=7, column=2, value="None - All constraints satisfied!")
        ok_cell.font = Font(bold=True, color="375623")
        ok_cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 65

    wb.save(filename)
    print("\nExcel file saved as:", filename)



#Running the code 
best,gen = genetic_algorithm()

print("\nBest Fitness:",fitness(best))
print("Generations Run:",gen)

show_solution(best)

print("\nRemaining Constraint Violations")

violations = violation_report(best)

if len(violations)==0:
    print("None")
else:
    for v in violations:
        print("-",v)

export_to_excel(best, fitness(best), gen)
